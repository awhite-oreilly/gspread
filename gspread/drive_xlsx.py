"""
gspread.drive_xlsx
~~~~~~~~~~~~~~~~~~

This module contains classes for reading .xlsx files from Google Drive.

"""

from io import BytesIO
from typing import Any, List, Optional, Union

try:
    from openpyxl import load_workbook
    from openpyxl.worksheet.worksheet import Worksheet as OpenpyxlWorksheet
    from openpyxl.workbook.workbook import Workbook as OpenpyxlWorkbook
except ImportError:
    raise ImportError(
        "openpyxl is required for reading .xlsx files. "
        "Install it with: pip install openpyxl"
    )

from .exceptions import ReadOnlyError, WorksheetNotFound
from .utils import GridRangeType, a1_to_rowcol, absolute_range_name
from .worksheet import ValueRange


class DriveXlsxSpreadsheet:
    """A read-only spreadsheet from a .xlsx file in Google Drive.

    This class provides a similar interface to Spreadsheet but reads
    from a downloaded .xlsx file instead of the Sheets API.
    """

    def __init__(self, file_id: str, file_name: str, file_content: bytes):
        """
        :param str file_id: Drive file ID
        :param str file_name: Name of the file in Drive
        :param bytes file_content: Downloaded .xlsx file content
        """
        self._id = file_id
        self._name = file_name

        try:
            self.workbook: OpenpyxlWorkbook = load_workbook(
                BytesIO(file_content), read_only=True, data_only=True
            )
        except Exception as e:
            from .exceptions import LocalFileError
            raise LocalFileError(f"Failed to parse .xlsx file: {e}")

    @property
    def id(self) -> str:
        """Spreadsheet ID (Drive file ID)."""
        return self._id

    @property
    def title(self) -> str:
        """Spreadsheet title (Drive file name)."""
        return self._name

    @property
    def url(self) -> str:
        """Spreadsheet URL (Drive file URL)."""
        return f"https://drive.google.com/file/d/{self._id}/view"

    @property
    def timezone(self) -> str:
        """.xlsx files don't have timezone info."""
        return "UTC"

    @property
    def locale(self) -> str:
        """.xlsx files don't have locale info."""
        return "en_US"

    @property
    def sheet1(self) -> "DriveXlsxWorksheet":
        """Shortcut property for getting the first worksheet."""
        return self.get_worksheet(0)

    def __iter__(self):
        yield from self.worksheets()

    def __repr__(self) -> str:
        return "<{} {} id:{}>".format(
            self.__class__.__name__,
            repr(self.title),
            self.id,
        )

    def worksheets(self, exclude_hidden: bool = False) -> List["DriveXlsxWorksheet"]:
        """Returns a list of all worksheets in the spreadsheet.

        :param exclude_hidden: If True, only return visible worksheets
        :type exclude_hidden: bool
        :returns: List of DriveXlsxWorksheet objects
        :rtype: list
        """
        worksheets = []
        for index, sheet_name in enumerate(self.workbook.sheetnames):
            sheet = self.workbook[sheet_name]
            if exclude_hidden and sheet.sheet_state == "hidden":
                continue
            worksheets.append(DriveXlsxWorksheet(self, sheet, index))
        return worksheets

    def get_worksheet(self, index: int) -> "DriveXlsxWorksheet":
        """Returns a worksheet with specified index.

        :param index: Worksheet index (starting from 0)
        :type index: int
        :returns: DriveXlsxWorksheet object
        :raises WorksheetNotFound: If index is out of range
        """
        try:
            sheet_name = self.workbook.sheetnames[index]
            sheet = self.workbook[sheet_name]
            return DriveXlsxWorksheet(self, sheet, index)
        except IndexError:
            raise WorksheetNotFound(f"index {index} not found")

    def worksheet(self, title: str) -> "DriveXlsxWorksheet":
        """Returns a worksheet with specified title.

        :param title: Worksheet title
        :type title: str
        :returns: DriveXlsxWorksheet object
        :raises WorksheetNotFound: If worksheet not found
        """
        if title not in self.workbook.sheetnames:
            raise WorksheetNotFound(title)

        index = self.workbook.sheetnames.index(title)
        sheet = self.workbook[title]
        return DriveXlsxWorksheet(self, sheet, index)

    # Write operations - raise ReadOnlyError
    def add_worksheet(self, *args, **kwargs):
        raise ReadOnlyError("Cannot add worksheets to a .xlsx file from Drive")

    def del_worksheet(self, *args, **kwargs):
        raise ReadOnlyError("Cannot delete worksheets from a .xlsx file from Drive")

    def batch_update(self, *args, **kwargs):
        raise ReadOnlyError("Cannot update a .xlsx file from Drive")


class DriveXlsxWorksheet:
    """A read-only worksheet from a .xlsx file in Google Drive.

    This class provides a similar interface to Worksheet but reads
    from a downloaded .xlsx file instead of the Sheets API.
    """

    def __init__(
        self,
        spreadsheet: DriveXlsxSpreadsheet,
        sheet: OpenpyxlWorksheet,
        index: int,
    ):
        self._spreadsheet = spreadsheet
        self._sheet = sheet
        self._index = index
        self.spreadsheet_id = spreadsheet.id

    @property
    def id(self) -> int:
        """Generate a numeric ID for the worksheet."""
        return hash(self._sheet.title) % (10**8)

    @property
    def spreadsheet(self) -> DriveXlsxSpreadsheet:
        """Parent spreadsheet."""
        return self._spreadsheet

    @property
    def title(self) -> str:
        """Worksheet title."""
        return self._sheet.title

    @property
    def index(self) -> int:
        """Worksheet index."""
        return self._index

    @property
    def url(self) -> str:
        """Worksheet URL."""
        return f"{self._spreadsheet.url}#gid={self.id}"

    @property
    def row_count(self) -> int:
        """Number of rows."""
        return self._sheet.max_row or 1

    @property
    def col_count(self) -> int:
        """Number of columns."""
        return self._sheet.max_column or 1

    @property
    def column_count(self) -> int:
        """Alias for col_count."""
        return self.col_count

    @property
    def frozen_row_count(self) -> int:
        """Number of frozen rows."""
        if self._sheet.freeze_panes:
            return self._sheet.freeze_panes.row - 1 if self._sheet.freeze_panes.row else 0
        return 0

    @property
    def frozen_col_count(self) -> int:
        """Number of frozen columns."""
        if self._sheet.freeze_panes:
            col_letter = self._sheet.freeze_panes.column
            if col_letter:
                return self._letter_to_index(col_letter) - 1
        return 0

    @property
    def isSheetHidden(self) -> bool:
        """Whether worksheet is hidden."""
        return self._sheet.sheet_state == "hidden"

    def __repr__(self) -> str:
        return "<{} {} id:{}>".format(
            self.__class__.__name__,
            repr(self.title),
            self.id,
        )

    @staticmethod
    def _letter_to_index(letter: str) -> int:
        """Convert column letter to index (A=1, B=2, etc.)."""
        result = 0
        for char in letter:
            result = result * 26 + (ord(char.upper()) - ord('A') + 1)
        return result

    def _read_range(
        self,
        start_row: int,
        start_col: int,
        end_row: Optional[int] = None,
        end_col: Optional[int] = None,
    ) -> List[List[Any]]:
        """Read a range of cells from the worksheet.

        :param start_row: Starting row (1-indexed)
        :param start_col: Starting column (1-indexed)
        :param end_row: Ending row (1-indexed, inclusive)
        :param end_col: Ending column (1-indexed, inclusive)
        :returns: List of lists containing cell values
        """
        if end_row is None:
            end_row = self.row_count
        if end_col is None:
            end_col = self.col_count

        values = []
        for row in self._sheet.iter_rows(
            min_row=start_row,
            max_row=end_row,
            min_col=start_col,
            max_col=end_col,
            values_only=True,
        ):
            # Convert None to empty string to match gspread behavior
            values.append([str(cell) if cell is not None else "" for cell in row])

        return values

    def get(
        self,
        range_name: Optional[str] = None,
        major_dimension: Optional[str] = None,
        value_render_option: Optional[str] = None,
        date_time_render_option: Optional[str] = None,
        combine_merged_cells: bool = False,
        maintain_size: bool = False,
        pad_values: bool = False,
        return_type: GridRangeType = GridRangeType.ValueRange,
    ) -> Union[ValueRange, List[List[str]]]:
        """Reads values from a range of cells.

        :param str range_name: Cell range in A1 notation (e.g., 'A1:B2').
            If not specified, returns all values.
        :returns: ValueRange or list of lists
        """
        if range_name is None:
            # Return all values
            values = self._read_range(1, 1, self.row_count, self.col_count)
        else:
            # Parse A1 notation
            if ":" in range_name:
                start, end = range_name.split(":")
                start_row, start_col = a1_to_rowcol(start)
                end_row, end_col = a1_to_rowcol(end)
            else:
                # Single cell
                start_row, start_col = a1_to_rowcol(range_name)
                end_row, end_col = start_row, start_col

            values = self._read_range(start_row, start_col, end_row, end_col)

        # Create ValueRange object to match gspread API
        if return_type == GridRangeType.ValueRange:
            value_range = ValueRange.from_json({
                "range": absolute_range_name(self.title, range_name or ""),
                "majorDimension": major_dimension or "ROWS",
                "values": values,
            })
            return value_range
        else:
            return values

    def get_all_values(self) -> List[List[str]]:
        """Returns all values in the worksheet as a list of lists.

        :returns: List of lists containing all cell values
        """
        return self._read_range(1, 1, self.row_count, self.col_count)

    def get_values(
        self, range_name: Optional[str] = None, **kwargs: Any
    ) -> List[List[str]]:
        """Alias for get() that returns list of lists."""
        result = self.get(range_name, return_type=GridRangeType.ListOfLists, **kwargs)
        return result if isinstance(result, list) else list(result)

    def row_values(
        self, row: int, value_render_option: Optional[str] = None
    ) -> List[str]:
        """Returns a list of all values in a row.

        :param int row: Row number (1-indexed)
        :returns: List of values
        """
        values = self._read_range(row, 1, row, self.col_count)
        return values[0] if values else []

    def col_values(
        self, col: int, value_render_option: Optional[str] = None
    ) -> List[str]:
        """Returns a list of all values in a column.

        :param int col: Column number (1-indexed)
        :returns: List of values
        """
        values = self._read_range(1, col, self.row_count, col)
        return [row[0] if row else "" for row in values]

    def get_all_records(
        self,
        empty2zero: bool = False,
        head: int = 1,
        default_blank: str = "",
        allow_underscores_in_numeric_literals: bool = False,
        numericise_ignore: Optional[List[Union[str, int]]] = None,
        value_render_option: Optional[str] = None,
    ) -> List[dict]:
        """Returns all records as a list of dictionaries.

        Uses the first row as keys.

        :param empty2zero: Whether to treat empty cells as zero
        :param head: Row number to use as header (1-indexed)
        :returns: List of dictionaries
        """
        from .utils import numericise_all, to_records

        all_values = self.get_all_values()
        if not all_values or len(all_values) < head:
            return []

        header_row = all_values[head - 1]
        data_rows = all_values[head:]

        if numericise_ignore is None:
            numericise_ignore = []

        # Convert column names to indices for numericise_ignore
        indices_to_ignore = []
        for item in numericise_ignore:
            if isinstance(item, str):
                try:
                    indices_to_ignore.append(header_row.index(item) + 1)
                except ValueError:
                    pass
            else:
                indices_to_ignore.append(item)

        # Numericise data rows
        numericised_rows = [
            numericise_all(
                row,
                empty2zero=empty2zero,
                default_blank=default_blank,
                allow_underscores_in_numeric_literals=allow_underscores_in_numeric_literals,
                ignore=indices_to_ignore,
            )
            for row in data_rows
        ]

        return to_records(header_row, numericised_rows)

    # Write operations - raise ReadOnlyError
    def update(self, *args, **kwargs):
        raise ReadOnlyError("Cannot update a .xlsx file from Drive")

    def batch_update(self, *args, **kwargs):
        raise ReadOnlyError("Cannot update a .xlsx file from Drive")

    def update_acell(self, *args, **kwargs):
        raise ReadOnlyError("Cannot update a .xlsx file from Drive")

    def update_cells(self, *args, **kwargs):
        raise ReadOnlyError("Cannot update a .xlsx file from Drive")

    def append_row(self, *args, **kwargs):
        raise ReadOnlyError("Cannot append to a .xlsx file from Drive")

    def append_rows(self, *args, **kwargs):
        raise ReadOnlyError("Cannot append to a .xlsx file from Drive")

    def insert_row(self, *args, **kwargs):
        raise ReadOnlyError("Cannot insert rows in a .xlsx file from Drive")

    def insert_rows(self, *args, **kwargs):
        raise ReadOnlyError("Cannot insert rows in a .xlsx file from Drive")

    def delete_rows(self, *args, **kwargs):
        raise ReadOnlyError("Cannot delete rows from a .xlsx file from Drive")

    def delete_columns(self, *args, **kwargs):
        raise ReadOnlyError("Cannot delete columns from a .xlsx file from Drive")

    def clear(self, *args, **kwargs):
        raise ReadOnlyError("Cannot clear a .xlsx file from Drive")

    def resize(self, *args, **kwargs):
        raise ReadOnlyError("Cannot resize a .xlsx file from Drive")

    def sort(self, *args, **kwargs):
        raise ReadOnlyError("Cannot sort a .xlsx file from Drive")
